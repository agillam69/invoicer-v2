"""Generate receipt documents in Excel format."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from invoice_manager.persistence.models import Invoice, Payment, Receipt


def generate_receipt_xlsx(
    receipt: Payment | Receipt,
    invoice: Invoice | None,
    settings: dict[str, Any],
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Receipt"
    sheet["A1"] = str(settings.get("business_name") or "Receipt")
    sheet["A1"].font = Font(size=16, bold=True)
    number = receipt.receipt_number if isinstance(receipt, Payment) else receipt.number
    sheet["A3"] = f"{settings.get('receipt_title') or 'RECEIPT'} — {number}"
    sheet["A3"].font = Font(size=14, bold=True)
    if invoice is not None:
        paid = sum(p.amount_cents for p in invoice.payments if not p.is_reversed)
        credits = sum(c.amount_cents for c in invoice.credits)
        rows: list[tuple[str, Any]] = [
            ("Received from", invoice.client_name),
            ("Invoice", invoice.number),
            ("Invoice date", invoice.issue_date),
            ("Invoice amount", invoice.total_cents / 100),
            ("Amount paid", receipt.amount_cents / 100),
            ("Payment method", receipt.method or ""),
            ("Reference", receipt.reference or ""),
            ("Amount outstanding", max(0, invoice.total_cents - paid - credits) / 100),
        ]
    else:
        assert isinstance(receipt, Receipt)
        rows = [
            ("Received from", receipt.client_name),
            ("Date", receipt.date),
            ("Amount received", receipt.amount_cents / 100),
            ("Payment method", receipt.method or ""),
            ("Reference", receipt.reference or ""),
            ("For", receipt.description or ""),
            ("Notes", receipt.notes or ""),
        ]
    for row, (label, value) in enumerate(rows, start=5):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        if "amount" in label.lower():
            sheet.cell(row=row, column=2).number_format = '$#,##0.00'
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 50
    workbook.save(output_path)
    return output_path
