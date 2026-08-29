"""Generate an Excel workbook version of an invoice."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from invoice_manager.domain.money import Money
from invoice_manager.domain.statuses import invoice_balance_cents
from invoice_manager.persistence.models import Invoice


def _get(settings: dict[str, Any], key: str, default: str = "") -> str:
    value = settings.get(key, default)
    if not value:
        return default
    return str(value)


def _fmt(cents: int) -> str:
    return Money(cents=cents).__str__()


def _write_cell(sheet: Any, row: int, col: int, value: Any, *, bold: bool = False, size: int = 11) -> None:
    cell = sheet.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size)


def generate_invoice_xlsx(invoice: Invoice, settings: dict[str, Any], output_path: Path) -> Path:
    """Create an .xlsx workbook for the invoice matching the reference layout."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet("Invoice")
    sheet.title = invoice.number

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)

    gst_rate = Decimal(_get(settings, "gst_rate", "0.0") or "0.0")
    doc_title = (
        _get(settings, "invoice_title_tax", "TAX INVOICE")
        if gst_rate > 0
        else _get(settings, "invoice_title", "INVOICE")
    )

    row = 1
    _write_cell(sheet, row, 1, doc_title, bold=True, size=14)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    # Business details (left) and invoice meta (right)
    paid_cents = sum(p.amount_cents for p in invoice.payments if not p.is_reversed)
    credit_cents = sum(c.amount_cents for c in invoice.credits)
    amount_paid = paid_cents + credit_cents
    balance_due = invoice_balance_cents(invoice)

    status = "Paid" if balance_due <= 0 else "Unpaid"
    if invoice.is_cancelled:
        status = "Cancelled"
    elif invoice.is_void:
        status = "Void"
    elif balance_due > 0 and invoice.due_date and cast(date, invoice.due_date) < date.today():
        status = "Overdue"

    meta_pairs = [
        ("Business Name", _get(settings, "business_name"), "Invoice Number", invoice.number),
        ("ABN", _get(settings, "business_abn"), "Invoice Date", str(invoice.issue_date)),
        ("Email", _get(settings, "business_email"), "Due Date", str(invoice.due_date or "")),
        ("Phone", _get(settings, "business_phone"), "Client Name", invoice.client_name),
        ("Address", _get(settings, "business_address"), "Client Status", status),
        (
            "Payment Terms",
            _get(settings, "invoice_payment_terms_note", "Payment due within {days} days").format(
                days=_get(settings, "payment_terms_days", "7")
            ),
            None,
            None,
        ),
    ]
    for left_label, left_value, right_label, right_value in meta_pairs:
        if left_label:
            _write_cell(sheet, row, 1, left_label, bold=True)
            _write_cell(sheet, row, 2, left_value)
        if right_label:
            _write_cell(sheet, row, 5, right_label, bold=True)
            _write_cell(sheet, row, 6, right_value)
        row += 1

    row += 1
    _write_cell(sheet, row, 1, "Bill To", bold=True)
    _write_cell(sheet, row, 2, invoice.client_name)
    row += 1
    _write_cell(sheet, row, 1, "Company", bold=True)
    _write_cell(sheet, row, 2, invoice.client_name)
    row += 1
    _write_cell(sheet, row, 1, "Address", bold=True)
    _write_cell(sheet, row, 2, invoice.client_address or "")
    row += 2

    headers = [
        _get(settings, "invoice_description_header", "Description"),
        _get(settings, "invoice_qty_header", "Qty"),
        _get(settings, "invoice_unit_header", "Unit"),
        _get(settings, "invoice_price_header", "Price"),
        _get(settings, "invoice_gst_header", "GST"),
        "Line Subtotal",
        "Line Total",
    ]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for item in invoice.items:
        sheet.cell(row=row, column=1, value=item.description)
        sheet.cell(row=row, column=2, value=item.quantity)
        sheet.cell(row=row, column=3, value=item.unit or "ea")
        sheet.cell(row=row, column=4, value=_fmt(item.unit_price_cents))
        sheet.cell(row=row, column=5, value=_fmt(item.gst_cents))
        sheet.cell(row=row, column=6, value=_fmt(item.subtotal_cents))
        sheet.cell(row=row, column=7, value=_fmt(item.total_cents))
        row += 1

    row += 1
    # Payment details on the left, summary on the right
    bank_name = _get(settings, "bank_name")
    bsb = _get(settings, "bank_bsb")
    account = _get(settings, "bank_account")
    account_name = _get(settings, "bank_account_name")

    summary_start = row
    _write_cell(sheet, row, 5, _get(settings, "invoice_subtotal_label", "Subtotal (ex GST)"), bold=True)
    sheet.cell(row=row, column=7, value=_fmt(invoice.subtotal_cents))
    row += 1
    _write_cell(sheet, row, 5, _get(settings, "invoice_gst_label", "GST"), bold=True)
    sheet.cell(row=row, column=7, value=_fmt(invoice.gst_cents))
    row += 1
    _write_cell(sheet, row, 5, _get(settings, "invoice_total_label", "Total (inc GST)"), bold=True)
    sheet.cell(row=row, column=7, value=_fmt(invoice.total_cents))
    row += 1
    _write_cell(sheet, row, 5, _get(settings, "invoice_amount_paid_label", "Amount Paid"), bold=True)
    sheet.cell(row=row, column=7, value=_fmt(amount_paid))
    row += 1
    _write_cell(sheet, row, 5, _get(settings, "invoice_balance_due_label", "Balance Due"), bold=True)
    sheet.cell(row=row, column=7, value=_fmt(balance_due))

    if bank_name or account:
        row = summary_start
        _write_cell(sheet, row, 1, _get(settings, "invoice_payment_details_label", "Payment Details"), bold=True, size=12)
        row += 1
        _write_cell(sheet, row, 1, "Bank Name", bold=True)
        sheet.cell(row=row, column=2, value=bank_name)
        row += 1
        _write_cell(sheet, row, 1, "Account Name", bold=True)
        sheet.cell(row=row, column=2, value=account_name)
        row += 1
        _write_cell(sheet, row, 1, "BSB", bold=True)
        sheet.cell(row=row, column=2, value=bsb)
        row += 1
        _write_cell(sheet, row, 1, "Account Number", bold=True)
        sheet.cell(row=row, column=2, value=account)
        row += 1
        _write_cell(sheet, row, 1, "Reference", bold=True)
        sheet.cell(row=row, column=2, value=f"Use {invoice.number} as reference")

    row = max(row, summary_start + 6)
    row += 2

    if invoice.notes:
        _write_cell(sheet, row, 1, _get(settings, "invoice_notes_label", "Notes:"), bold=True)
        sheet.cell(row=row, column=2, value=invoice.notes)
        row += 1

    gst_footer = _get(settings, "invoice_gst_footer_note", "")
    if gst_footer:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = sheet.cell(row=row, column=1, value=gst_footer)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 2

    thank_you = _get(settings, "invoice_thank_you", "Thank you for your business!")
    if thank_you:
        _write_cell(sheet, row, 1, thank_you, bold=True)

    for col in range(1, 8):
        sheet.column_dimensions[chr(64 + col)].auto_size = True

    workbook.save(output_path)
    return output_path
